from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.report.db_manager import save_recommendations, update_prices

OUTPUT_FILE = (
    Path(__file__).parent.parent.parent / "data" / "reports" / "株式投資推奨レポート.xlsx"
)

C_HEADER  = "1F4E79"
C_SUBHEAD = "2E75B6"
C_TIME    = "243F60"
C_LEGEND  = "F0F4FA"
C_RANK1   = "FFF2CC"
C_RANK2   = "DEEAF1"
C_RANK3   = "E2EFDA"
C_DEFAULT = "F5F5F5"
RANK_COLORS = {1: C_RANK1, 2: C_RANK2, 3: C_RANK3}

FLOW_LABELS = {
    "ニュース起点":       "ニュース分析",
    "YouTube起点":        "YouTube動画分析",
    "スクリーニング起点": "条件スクリーニング",
}

LEGEND_ROWS = [
    ("おすすめ度",          "★〜★★★★★",      "★が多いほど強くおすすめ。★★★★★=強く買い推奨、★★=様子見、★=要注意"),
    ("AIの確信度(%)",       "0〜100%",          "AIが「この銘柄はよい」と判断した確かさ。80%以上が高信頼"),
    ("売買タイミング(RSI)", "0〜100",           "30以下=売られすぎで反発チャンス、70以上=買われすぎで注意、30〜50=スイング買い好機"),
    ("トレンド方向(MACD)",  "上昇/下落/横ばい", "上昇=今から勢いが出てくる、下落=まだ下がり続ける可能性"),
    ("20日平均株価比",      "例: -5.2%",        "直近20日間の平均株価と比べて今の株価が何%高い/安いか。マイナスは平均より安い"),
    ("価格帯の位置(BB)",    "下限/中間/上限",   "「ボリンジャーバンド」内での位置。下限付近=割安ゾーン、上限付近=割高ゾーン"),
    ("短期過熱度(Stoch)",   "0〜100",           "20以下=売られすぎ(買いチャンス)、80以上=買われすぎ(利確タイミング)"),
    ("スイングスコア",      "0〜10点",          "スイング取引（数日〜2週間保有）向けの総合点。7点以上が買い候補"),
]


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value, bold=False, bg=None, fg="000000",
          align="left", size=10, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=fg, size=size, name="游ゴシック")
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border    = _border()
    if bg:
        c.fill = _fill(bg)
    return c


def _last_row(ws) -> int:
    return ws.max_row if ws.max_row and ws.max_row > 1 else 0


def _write_legend_sheet(wb):
    if "指標の説明" in wb.sheetnames:
        return
    ws = wb.create_sheet(title="指標の説明")

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value     = "各項目の見方・指標の説明"
    c.font      = Font(bold=True, color="FFFFFF", size=14, name="游ゴシック")
    c.fill      = _fill(C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")
    c = ws["A2"]
    c.value     = "※ 投資は自己責任です。このレポートはAIによる分析であり、投資の結果を保証するものではありません。"
    c.font      = Font(italic=True, color="C00000", size=9, name="游ゴシック")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    for col, h in enumerate(["項目名", "表示形式の例", "見方・意味"], 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=10, name="游ゴシック")
        c.fill      = _fill(C_SUBHEAD)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()
    ws.row_dimensions[3].height = 18

    for i, (name, fmt, desc) in enumerate(LEGEND_ROWS, 4):
        bg = C_LEGEND if i % 2 == 0 else "FFFFFF"
        for col, val in enumerate([name, fmt, desc], 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font      = Font(size=10, name="游ゴシック")
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border    = _border()
            c.fill      = _fill(bg)
        ws.row_dimensions[i].height = 36

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 62

    stars_row = len(LEGEND_ROWS) + 5
    ws.merge_cells(f"A{stars_row}:C{stars_row}")
    c = ws.cell(row=stars_row, column=1, value="【おすすめ度の目安】")
    c.font      = Font(bold=True, color="FFFFFF", size=10, name="游ゴシック")
    c.fill      = _fill(C_SUBHEAD)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[stars_row].height = 18

    stars_data = [
        ("★★★★★", C_RANK1, "強くおすすめ — 複数の指標が揃った買いチャンス"),
        ("★★★★☆", C_RANK2, "おすすめ — 条件の多くが揃っている"),
        ("★★★☆☆", C_RANK3, "やや買い — いくつかの条件が揃っている"),
        ("★★☆☆☆", "FFFFFF", "様子見 — 条件が弱い、追加確認が必要"),
        ("★☆☆☆☆", "FFFFFF", "要注意 — リスクが高い、または根拠が弱い"),
    ]
    for j, (stars, bg, meaning) in enumerate(stars_data, stars_row + 1):
        ws.merge_cells(f"A{j}:B{j}")
        c = ws.cell(row=j, column=1, value=stars)
        c.font      = Font(bold=True, size=11, name="游ゴシック")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()
        c.fill      = _fill(bg)
        c2 = ws.cell(row=j, column=3, value=meaning)
        c2.font      = Font(size=10, name="游ゴシック")
        c2.alignment = Alignment(horizontal="left", vertical="center")
        c2.border    = _border()
        c2.fill      = _fill(bg)
        ws.row_dimensions[j].height = 20


def _write_block(ws, result: dict, start_row: int, now: datetime):
    row = start_row

    flow_raw   = result.get("flow", "")
    flow_label = FLOW_LABELS.get(flow_raw, flow_raw)
    has_news_col = flow_raw in ("ニュース起点", "YouTube起点")
    N = 13 if has_news_col else 12

    # ── 実行時刻ヘッダー ────────────────────────
    vix = result.get("vix")
    fg  = result.get("fear_greed")
    sentiment_str = ""
    if vix:
        sentiment_str += f"  VIX:{vix}"
    if fg:
        sentiment_str += f"  恐怖&欲指数:{fg}/100"
    ws.merge_cells(f"A{row}:{get_column_letter(N)}{row}")
    c = ws.cell(row=row, column=1,
                value=f"実行: {now.strftime('%H:%M')}　【{flow_label}】　{result.get('market','')}{sentiment_str}")
    c.font      = Font(bold=True, color="FFFFFF", size=11, name="游ゴシック")
    c.fill      = _fill(C_TIME)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    # ── 列ヘッダー ──────────────────────────────
    headers = [
        "順位",
        "コード",
        "銘柄名",
        "おすすめ度",
        "AIの\n確信度(%)",
        "株価\n(終値)",
        "売買タイミング\n(RSI)",
        "トレンド方向\n(MACD)",
        "20日平均\n株価比",
        "価格帯の\n位置(BB)",
        "短期過熱度\n(Stoch)",
        "おすすめ理由と売買タイミング",
    ]
    if has_news_col:
        headers.append("参考にしたニュース・動画")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=9, name="游ゴシック")
        c.fill      = _fill(C_SUBHEAD)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
    ws.row_dimensions[row].height = 32
    row += 1

    # ── ランキング ──────────────────────────────
    for item in result.get("rankings", []):
        rank = item.get("rank", row)
        bg   = RANK_COLORS.get(rank, C_DEFAULT)
        _cell(ws, row, 1,  rank,                       bold=True, bg=bg, align="center")
        _cell(ws, row, 2,  item.get("ticker", ""),                bg=bg, align="center")
        _cell(ws, row, 3,  item.get("name", ""),                  bg=bg)
        _cell(ws, row, 4,  item.get("stars", ""),                 bg=bg, align="center")
        _cell(ws, row, 5,  item.get("confidence"),                bg=bg, align="center")
        _cell(ws, row, 6,  item.get("close"),                     bg=bg, align="right")
        _cell(ws, row, 7,  item.get("RSI14"),                     bg=bg, align="center")
        _cell(ws, row, 8,  item.get("MACD方向", ""),              bg=bg, align="center")
        _cell(ws, row, 9,  item.get("SMA20比", ""),               bg=bg, align="center")
        _cell(ws, row, 10, item.get("BB位置", ""),                bg=bg, align="center")
        _cell(ws, row, 11, item.get("STOCH_K"),                   bg=bg, align="center")
        _cell(ws, row, 12, item.get("reason", ""),                bg=bg, wrap=True)
        if has_news_col:
            _cell(ws, row, 13, item.get("news_basis", ""),        bg=bg, wrap=True)
        ws.row_dimensions[row].height = 45
        row += 1

    # ── 総評 ────────────────────────────────────
    ws.merge_cells(f"A{row}:{get_column_letter(N)}{row}")
    c = ws.cell(row=row, column=1, value="【まとめ・市場の状況】")
    c.font      = Font(bold=True, color="FFFFFF", size=10, name="游ゴシック")
    c.fill      = _fill(C_SUBHEAD)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 16
    row += 1

    summary_text = result.get("summary", "")
    if result.get("market_outlook"):
        summary_text += f"\n【市場全体の状況】{result['market_outlook']}"
    ws.merge_cells(f"A{row}:{get_column_letter(N)}{row}")
    c = ws.cell(row=row, column=1, value=summary_text)
    c.font      = Font(size=10, name="游ゴシック")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border    = _border()
    ws.row_dimensions[row].height = 65
    row += 1

    return row


def export(result: dict) -> Path:
    """
    結果を 株式投資推奨レポート.xlsx に書き込む。
    - シート名: YYYYMMDD（1日1シート）
    - 同日に複数回実行した場合は同シートに追記
    """
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    now        = datetime.now()
    sheet_name = now.strftime("%Y%m%d")

    if OUTPUT_FILE.exists():
        wb = openpyxl.load_workbook(OUTPUT_FILE)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    _write_legend_sheet(wb)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        start_row = _last_row(ws) + 2
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.merge_cells("A1:M1")
        c = ws["A1"]
        c.value     = f"株式投資おすすめレポート　{now.strftime('%Y年%m月%d日')}"
        c.font      = Font(bold=True, color="FFFFFF", size=14, name="游ゴシック")
        c.fill      = _fill(C_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        # 列幅: 順位,コード,銘柄名,おすすめ度,確信度,株価,RSI,MACD,SMA20比,BB位置,Stoch,理由,ニュース
        for col, w in enumerate([5, 11, 18, 10, 9, 10, 10, 10, 9, 9, 9, 48, 52], 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = "A3"
        start_row = 2

    _write_block(ws, result, start_row, now)

    save_recommendations(result)
    update_prices()

    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE
