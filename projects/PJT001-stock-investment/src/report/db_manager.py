"""
推奨履歴データベース管理

Excel の「DB_推奨履歴」シートに全推奨記録を蓄積し、
実行ごとに現在価格・損益%を自動更新する。
"""
from datetime import datetime, date
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import yfinance as yf
import pandas as pd

DB_SHEET = "DB_推奨履歴"

# DB列定義
DB_COLS = [
    "推奨日", "推奨時刻", "フロー", "銘柄コード", "銘柄名",
    "推奨順位", "推奨度", "推奨時終値", "現在価格", "損益%", "最終更新",
    "RSI14", "MACD方向", "SMA20比", "BB位置",
]

OUTPUT_FILE = (
    Path(__file__).parent.parent.parent / "data" / "reports" / "株式投資推奨レポート.xlsx"
)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _init_db_sheet(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    """DB シートを新規作成してヘッダーを書く"""
    ws = wb.create_sheet(title=DB_SHEET)

    # タイトル
    ws.merge_cells(f"A1:{get_column_letter(len(DB_COLS))}1")
    c = ws["A1"]
    c.value     = "推奨履歴データベース"
    c.font      = Font(bold=True, color="FFFFFF", size=13, name="游ゴシック")
    c.fill      = _fill("243F60")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # ヘッダー
    for col, h in enumerate(DB_COLS, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=10, name="游ゴシック")
        c.fill      = _fill("2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()

    # 列幅
    widths = [12, 8, 14, 12, 16, 8, 10, 12, 12, 8, 12, 8, 10, 10, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A3"
    return ws


def _get_or_create_db(wb: openpyxl.Workbook):
    if DB_SHEET in wb.sheetnames:
        return wb[DB_SHEET]
    return _init_db_sheet(wb)


def save_recommendations(result: dict, wb: openpyxl.Workbook):
    """推奨結果を DB シートに追記する"""
    ws  = _get_or_create_db(wb)
    now = datetime.now()

    for item in result.get("rankings", []):
        next_row = ws.max_row + 1
        row_data = [
            str(date.today()),
            now.strftime("%H:%M"),
            result.get("flow", ""),
            item.get("ticker", ""),
            item.get("name", ""),
            item.get("rank"),
            item.get("stars", ""),
            item.get("close"),
            None,   # 現在価格（後で更新）
            None,   # 損益%
            None,   # 最終更新
            item.get("RSI14"),
            item.get("MACD方向", ""),
            item.get("SMA20比", ""),
            item.get("BB位置", ""),
        ]
        # 奇数/偶数行で背景色を交互に
        bg = "EBF3FB" if next_row % 2 == 0 else "FFFFFF"
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=next_row, column=col, value=val)
            c.font      = Font(size=10, name="游ゴシック")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = _border()
            c.fill      = _fill(bg)
        ws.row_dimensions[next_row].height = 18


def update_prices(wb: openpyxl.Workbook):
    """DB シートの現在価格・損益%・最終更新日を一括更新する"""
    if DB_SHEET not in wb.sheetnames:
        return
    ws = wb[DB_SHEET]

    # 列インデックスをマッピング
    col_map = {ws.cell(row=2, column=c).value: c for c in range(1, len(DB_COLS) + 1)}
    c_ticker   = col_map.get("銘柄コード", 4)
    c_entry    = col_map.get("推奨時終値", 8)
    c_current  = col_map.get("現在価格", 9)
    c_pnl      = col_map.get("損益%", 10)
    c_updated  = col_map.get("最終更新", 11)

    # ユニークな ticker を収集して一括取得
    tickers = set()
    for row in range(3, ws.max_row + 1):
        t = ws.cell(row=row, column=c_ticker).value
        if t:
            tickers.add(t)

    prices = {}
    for t in tickers:
        try:
            info = yf.Ticker(t).fast_info
            prices[t] = round(float(info.last_price), 2)
        except Exception:
            pass

    today = str(date.today())
    for row in range(3, ws.max_row + 1):
        ticker = ws.cell(row=row, column=c_ticker).value
        entry  = ws.cell(row=row, column=c_entry).value
        if ticker and ticker in prices:
            current = prices[ticker]
            ws.cell(row=row, column=c_current).value  = current
            ws.cell(row=row, column=c_updated).value  = today
            if entry and entry > 0:
                pnl = round((current - entry) / entry * 100, 2)
                c   = ws.cell(row=row, column=c_pnl)
                c.value = pnl
                # 損益の色分け
                c.font = Font(
                    size=10, name="游ゴシック",
                    color="C00000" if pnl < 0 else "375623",
                    bold=True,
                )


def load_performance_summary() -> str:
    """
    過去推奨の実績サマリーを文字列で返す。
    Claude のプロンプトに組み込んで精度向上に使う。
    """
    if not OUTPUT_FILE.exists():
        return ""
    try:
        wb = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)
        if DB_SHEET not in wb.sheetnames:
            return ""
        ws = wb[DB_SHEET]

        col_map = {ws.cell(row=2, column=c).value: c for c in range(1, len(DB_COLS) + 1)}
        rows = []
        for row in range(3, ws.max_row + 1):
            r = {h: ws.cell(row=row, column=col_map[h]).value for h in DB_COLS if h in col_map}
            if r.get("損益%") is not None:
                rows.append(r)

        if not rows:
            return ""

        df = pd.DataFrame(rows)
        df["損益%"] = pd.to_numeric(df["損益%"], errors="coerce")

        lines = [f"【過去推奨の実績】（{len(df)}件）"]

        # 全体勝率
        wins = (df["損益%"] > 0).sum()
        avg  = df["損益%"].mean()
        lines.append(f"  全体: 勝率 {wins}/{len(df)} ({wins/len(df)*100:.0f}%)  平均損益 {avg:+.1f}%")

        # 条件別実績
        for macd in ["買い", "売り"]:
            sub = df[df["MACD方向"] == macd]
            if len(sub) >= 2:
                lines.append(f"  MACD={macd}: 平均 {sub['損益%'].mean():+.1f}% ({len(sub)}件)")

        for rsi_label, rsi_cond in [("RSI<35", df["RSI14"] < 35), ("RSI35-50", (df["RSI14"] >= 35) & (df["RSI14"] < 50))]:
            sub = df[rsi_cond]
            if len(sub) >= 2:
                lines.append(f"  {rsi_label}: 平均 {sub['損益%'].mean():+.1f}% ({len(sub)}件)")

        for bb in ["下限付近", "中央付近", "上限付近"]:
            sub = df[df["BB位置"] == bb]
            if len(sub) >= 2:
                lines.append(f"  BB={bb}: 平均 {sub['損益%'].mean():+.1f}% ({len(sub)}件)")

        return "\n".join(lines)

    except Exception:
        return ""
